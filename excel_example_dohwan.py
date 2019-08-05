import collections
from openpyxl import load_workbook
import pandas as pd

# Output all contents of file.
excel_file = pd.read_excel('test.xlsx', sheet_name='Sheet1')
print(excel_file)

load_wb = load_workbook('test.xlsx', data_only=False)    # You must "data_only = True". If not this, you are see function value
load_ws = load_wb['Sheet1']                             # Enter Sheet Name
rows = load_ws.max_row + 1                              # Number of rows (add title rows)
cols = load_ws.max_column + 1                           # Number of colmn (add title rows)

count_list = []
Percent_list = []

for k in range(2,cols):
    value_list = []
    for j in range(1, rows):
        if j == 1 :
            print(load_ws.cell(j, k).value)
        else :   
            value_list.append(load_ws.cell(j, k).value)
        
    value_list = set(value_list)
    print(value_list)