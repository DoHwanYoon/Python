import collections
from openpyxl import load_workbook
import pandas as pd

# Output all contents of file.
excel_file = pd.read_excel('test.xlsx', sheet_name='Sheet1')

print(excel_file)

load_wb = load_workbook('test.xlsx', data_only=False)    # You must "data_only = True". If not this, you are see function value
load_ws = load_wb['Sheet1']                             # Enter Sheet Name
rows = load_ws.max_row + 1                              # Number of rows (add title rows)
cols = load_ws.max_column + 1                           # Number of colmn (add title rows)

append_list=[]
cnt_0 =0
cnt_1=0
data_list_0 =[] # 결과 리스트의 [0]밸류값을 가지는 데이터 
data_list_1 =[] # 결과 리스트의 [0]밸류값을 가지는 데이터
result_list = [] #결과 인덱스 밸류
result_count_dic = {} # result_list 밸류의 카운트 딕셔너리
value_list = [] #각 인덱스별 밸류 종류
bayes_value_dic_0 = {} # data_list_0 밸류의 카운트 딕셔너리
bayes_value_dic_1 = {} # data_list_1 밸류의 카운트 딕셔너리
bayes_result_dic = {} # 결과 밸류의 카운트 딕셔너리

# 결과 인덱스에 대한 value값 종류 추출
for k in range(2,cols-1): 
        for j in range(2, rows):
                append_list.append(load_ws.cell(j,k).value)
value_list = list(set(append_list))
append_list=[]

# 결과 value list
for k in range(2, rows):
        append_list.append(load_ws.cell(k, cols-1).value)
result_list = list(set(append_list))

# 결과 인덱스별 데이터
for k in range(2,rows): 
    append_list_0 = []
    append_list_1 =[]
    if load_ws.cell(k,cols-1).value == result_list[0] :
        cnt_0 += 1
        for i in range(2, cols):
            append_list_0.append(load_ws.cell(k, i).value)
        data_list_0.append(append_list_0)
        
    elif load_ws.cell(k,cols-1).value == result_list[1]: # 결과 밸류값이 두개라면 else 사용 가능.
        cnt_1 += 1
        for i in range(2, cols):
            append_list_1.append(load_ws.cell(k, i).value)
        data_list_1.append(append_list_1)
result_count_dic[result_list[0]] = cnt_0
result_count_dic[result_list[1]] = cnt_1

for i in range(0,2):
    bayes_result_dic[result_list[i]] = float(result_count_dic[result_list[i]]) / float(rows-2)

# 각 결과에 대한 인덱스별 밸류 counting
for i in value_list :
    cnt_0 = 0
    cnt_1 = 0
    for j in range(len(data_list_0)):
        for k in range(len(data_list_0[j])-1):
            if(i == data_list_0[j][k]):
                cnt_0 += 1
                bayes_value_dic_0[i] = float(cnt_0) / float(result_count_dic[result_list[0]]) 
                
    for j in range(len(data_list_1)):
        for k in range(len(data_list_1[j])-1):
            if(i == data_list_1[j][k]):
                cnt_1 += 1
                bayes_value_dic_1[i] = float(cnt_1) / float(result_count_dic[result_list[1]])

# 각 결과별로 없는 인덱스 값에 대해 추가               
for i in value_list:
    a = i in bayes_value_dic_0
    if a == False:
        bayes_value_dic_0[i] = 0.0

# 각 결과별로 없는 인덱스 값에 대해 추가
for i in value_list:
    a = i in bayes_value_dic_1
    if a == False:
        bayes_value_dic_1[i] = 0.0

'''
# 확률 출력
for key, val in bayes_value_dic_0.items():
    print(key+"|"+result_list[0],"=",val,"/",result_count_dic[result_list[0]],"  확률 :", float(val)/float(result_count_dic[result_list[0]]))
print()
for key, val in bayes_value_dic_1.items():
    print(key+"|"+result_list[1],"=",val,"/",result_count_dic[result_list[1]],"  확률 :", float(val)/float(result_count_dic[result_list[1]]))
'''
print()
print("결과 확률")
print(bayes_result_dic)
print(result_list[0]+"의 확률")
print(bayes_value_dic_0)
print(result_list[1]+"의 확률")
print(bayes_value_dic_1)

# Output all contents of file.
test_data = pd.read_excel('test_data.xlsx', sheet_name='Sheet1')
#print(excel_file)

test_data_wb = load_workbook('test_data.xlsx', data_only=False)    # You must "data_only = True". If not this, you are see function value
test_data_ws = test_data_wb['Sheet1']                             # Enter Sheet Name
data_rows = test_data_ws.max_row + 1                          # Number of rows (add title rows)
data_cols = test_data_ws.max_column + 1

print()
print("결과")
# 확률계산
for j in range(2,data_rows):
    result_0 = 1.0
    result_1 = 1.0
    for i in range(2, data_cols):
        test_d = test_data_ws.cell(j, i).value in bayes_value_dic_0
        if test_d == True:
            #print(test_data_ws.cell(2, i).value)
            result_0 *= bayes_value_dic_0[test_data_ws.cell(j, i).value]
    print(test_data_ws.cell(j, 1).value ,"|", result_list[0],"=",result_0*result_count_dic[result_list[0]])

    for k in range(2, data_cols):
        test_d = test_data_ws.cell(j, k).value in bayes_value_dic_1
        if test_d == True:
            #print(test_data_ws.cell(2, i).value)
            result_1 *= bayes_value_dic_1[test_data_ws.cell(j, k).value]
    print(test_data_ws.cell(j, 1).value ,"|", result_list[1],"=",result_1*result_count_dic[result_list[1]])
    
'''
for j in range(2, data_rows):     
    for i in range(2, data_cols):
        test_d = test_data_ws.cell(j, i).value in bayes_value_dic_1
        if test_d == True:
            #print(test_data_ws.cell(2, i).value)
            result_1 *= bayes_value_dic_1[test_data_ws.cell(j, i).value]
    print(test_data_ws.cell(j, 1).value ,"|", result_list[1],"=",result_1*result_count_dic[result_list[1]])
    
'''
'''
for i in range(2,data_rows):
    print(test_data_ws.cell(i, 1).value ,"|", result_list[0],"=",result_0*result_count_dic[result_list[0]])
    print(test_data_ws.cell(i, 1).value ,"|", result_list[1],"=",result_1*result_count_dic[result_list[1]])
    print()
'''
