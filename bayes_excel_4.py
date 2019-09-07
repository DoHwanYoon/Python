giimport collections
from openpyxl import load_workbook
import pandas as pd

# Output all contents of file.
excel_file = pd.read_excel('test.xlsx', sheet_name='Sheet1')

load_wb = load_workbook('test.xlsx', data_only=False)    # You must "data_only = True". If not this, you are see function value
load_ws = load_wb['Sheet1']                             # Enter Sheet Name
rows = load_ws.max_row + 1                              # Number of rows (add title rows)
cols = load_ws.max_column + 1                           # Number of colmn (add title rows)

append_list=[]
yes_data_list =[] # yes값을 가지는 데이터 
no_data_list =[] # no값을 가지는 데이터
result_list = [] #결과 인덱스 밸류
value_list = [] #각 인덱스별 밸류 종류
yes_count_list = []  #
no_count_list = []
result_count_list =[] #
value_count_list =[]
bayes_value_dic = {}

# 결과 인덱스에 대한 value값 종류 추출
for k in range(2,cols): 
    append_list = []
    for j in range(2, rows):
        append_list.append(load_ws.cell(j,k).value)
    append_list = set(append_list)
    value_list.append([append_list])
value_list = list(value_list)

# 결과 인덱스별 데이터
for k in range(2,rows): 
        yes_append_list = []
        no_append_list =[]
        if load_ws.cell(k,cols-1).value == value_list[0] :
                for i in range(2, cols):
                        yes_append_list.append(load_ws.cell(k, i).value)
                yes_data_list.append(yes_append_list)
        
        else:
                for i in range(2, cols):
                        no_append_list.append(load_ws.cell(k, i).value)
                no_data_list.append(no_append_list)

for i in yes_data_list :    # 각 밸류의 개수 count
    print(i)
        








#print(yes_data_list)
#print(no_data_list)           
            
            
#print(result_list)
#print(yes_data_list)
#print(no_data_list)
