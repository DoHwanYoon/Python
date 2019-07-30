'''
    2019 인제대학교 영재교육원 정보과학반 사사과정
    Excel 파일 입출력 및 원하는 값에 따른 출력
    작성자 : 2019 정보과학반 조교 당현아
'''

from openpyxl import load_workbook
import pandas as pd

find_column = 0

# Output all contents of file.
excel_file = pd.read_excel('test.xlsx', sheet_name='Sheet1')
print(excel_file)

load_wb = load_workbook('test.xlsx', data_only=True)    # You must "data_only = True". If not this, you are see function value
load_ws = load_wb['Sheet1']                             # Enter Sheet Name
rows = load_ws.max_row+1                                  # Number of rows (add title rows)
cols = load_ws.max_column+1                             # Number of colmn (add title rows)

# Enter the name of the column you want.
column_name = input("찾고자하는 열의 이름을 쓰세요 : ")

# Finds a column with the same name as column_name.
for i in range(1, cols):
    if column_name == load_ws.cell(1, i).value :
        find_column = i

# Outputs all values in the corresponding columns.
for i in range(2, rows):
    print(load_ws.cell(i, find_column).value)